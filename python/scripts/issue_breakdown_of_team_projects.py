import os
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

# Load API key from environment
load_dotenv()
FOSSA_API_KEY = os.getenv("FOSSA_API_KEY")
BASE_URL = "https://app.fossa.com/api"

# Fetch team name
def fetch_team_name(team_id):
    response = requests.get(
        f"{BASE_URL}/teams/{team_id}",
        headers={"Authorization": f"Bearer {FOSSA_API_KEY}"}
    )
    response.raise_for_status()
    return response.json()["name"]

# Fetch issues by category
def fetch_issues_by_category(team_id, category):
    issues, page = [], 1
    while True:
        response = requests.get(
            f"{BASE_URL}/v2/issues",
            params={
                "category": category,
                "status": "active",
                "scope[type]": "global",
                "teamId[0]": team_id,
                "sort": "created_at_desc",
                "page": page,
                "count": 1000
            },
            headers={"Authorization": f"Bearer {FOSSA_API_KEY}"}
        )
        response.raise_for_status()
        data = response.json()
        issues.extend(data["issues"])
        if len(data["issues"]) < 1000:
            break
        page += 1
    return issues

# Summarize issues by severity or type
def summarize_issues(issues, category):
    if category == "licensing":
        return {
            "Denied": sum(1 for issue in issues if issue["type"] == "policy_conflict"),
            "Flagged": sum(1 for issue in issues if issue["type"] == "policy_flag"),
            "Unlicensed": sum(1 for issue in issues if issue["type"] == "unlicensed")
        }
    return {
        severity: sum(1 for issue in issues if issue.get("severity") == severity)
        for severity in ["Critical", "High", "Medium", "Low", "Unknown"]
    }

# Generate Excel report
def generate_excel_report(team_name, vulnerability_summary, license_summary, quality_summary):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary Report"
    worksheet["A1"], worksheet["A1"].font = f"{team_name} Summary Report", Font(bold=True, size=14)
    
    row = 3  # Start below title
    
    # Security Findings
    row = write_summary(worksheet, "Security Findings", vulnerability_summary, row)
    row = write_summary(worksheet, "OSS License Issues", license_summary, row)
    write_summary(worksheet, "Quality Issues", quality_summary, row)
    
    workbook.save("CWP_Summary_Report.xlsx")
    print("Report generated successfully: CWP_Summary_Report.xlsx")

# Helper to write section with title and summary to Excel
def write_summary(ws, title, summary, start_row):
    ws[f"A{start_row}"], ws[f"A{start_row}"].font = title, Font(bold=True)
    start_row += 1
    ws[f"A{start_row}"], ws[f"B{start_row}"] = "Type", "Nr. of Findings"
    ws[f"A{start_row}"].font = ws[f"B{start_row}"].font = Font(bold=True)
    start_row += 1

    for item, count in summary.items():
        ws[f"A{start_row}"], ws[f"B{start_row}"] = item, count
        start_row += 1
    return start_row + 1  # Add a blank row before the next section

# Main function to fetch and process data
def main(team_id):
    team_name = fetch_team_name(team_id)
    vulnerability_issues = fetch_issues_by_category(team_id, "vulnerability")
    license_issues = fetch_issues_by_category(team_id, "licensing")
    quality_issues = fetch_issues_by_category(team_id, "quality")

    vulnerability_summary = summarize_issues(vulnerability_issues, "vulnerability")
    license_summary = summarize_issues(license_issues, "licensing")
    quality_summary = summarize_issues(quality_issues, "quality")

    generate_excel_report(team_name, vulnerability_summary, license_summary, quality_summary)

if __name__ == "__main__":
    team_id = 55413  # Replace this with your team ID or make it a command-line argument
    main(team_id)
